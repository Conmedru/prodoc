"""Excel exporter for scraped doctor data."""
import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from scraper.parsers import DoctorInfo
from scraper.config import OUTPUT_DIR

logger = logging.getLogger(__name__)

COLUMNS = [
    ("№", 5),
    ("ФИО", 35),
    ("Город", 18),
    ("Специальность", 35),
    ("Место работы", 45),
    ("Адрес работы", 45),
    ("Стаж", 15),
    ("URL", 50),
]


def export_to_excel(doctors: list[DoctorInfo], filename: str | None = None) -> str:
    """Export doctor data to Excel file.
    
    Returns:
        Path to the created Excel file.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    if filename is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"doctors_{ts}.xlsx"
    
    filepath = os.path.join(OUTPUT_DIR, filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Врачи"
    
    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # Headers
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = col_width
    
    # Freeze top row
    ws.freeze_panes = "A2"
    
    # Data rows
    row_num = 2
    for idx, doc in enumerate(doctors, 1):
        specialties_str = ", ".join(doc.specialties)
        
        # Multiple workplaces -> one per line
        workplaces_str = "\n".join(doc.current_workplaces) if doc.current_workplaces else ""
        addresses_str = "\n".join(doc.work_addresses) if doc.work_addresses else ""
        
        row_data = [
            idx,
            doc.full_name,
            doc.city,
            specialties_str,
            workplaces_str,
            addresses_str,
            doc.experience_years,
            doc.url,
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
        
        row_num += 1
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:H{row_num - 1}"
    
    wb.save(filepath)
    logger.info(f"Exported {len(doctors)} doctors to {filepath}")
    return filepath
